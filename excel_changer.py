# -*- coding: utf-8 -*-
"""
Created on Tue Mar 14 23:49:18 2017

@author: wtw
"""

from openpyxl import load_workbook
import numpy as np

# read the data
wb = load_workbook('input.xlsx')
ws = wb.get_sheet_by_name('Data')

# now read the data one by one
max_num = 10000
starting_num = 1
data_dict = {}
row_id = np.array(range(max_num)) * 4 + 1
# read the country list
for i_data in row_id:
    country_item = ws['A' + str(i_data)]
    
    # check if the sheet has run out    
    if country_item.data_type == 'n':
        break
    
    country_name = country_item.value
    country_dict = {}
    
    # process the data one by one:
    country_dict['name'] = country_name
    country_dict['code'] = ws['B' + str(i_data)].value
    country_dict['i_1'] = ws['D' + str(i_data + 1)].value
    country_dict['i_2'] = ws['D' + str(i_data + 2)].value
    country_dict['i_3'] = ws['D' + str(i_data + 3)].value
    country_dict['i_4'] = ws['D' + str(i_data + 4)].value
    data_dict[country_dict['name']] = country_dict
    
# create a new sheet
ws = wb.create_sheet('new_data')

# init the column
ws['A1'] = 'Country Name'
ws['B1'] = 'Country Code'
ws['C1'] = 'Foreign direct investment, net inflows (BoP, current US$)'
ws['D1'] = 'GNI per capita, PPP (current international $)'
ws['E1'] = 'Population, total'
ws['F1'] = 'Urban population'

row_counter = 1
for country_name in data_dict:
    row_counter = row_counter + 1
    country_dict = data_dict[country_name]
    ws['A' + str(row_counter)] = country_dict['code']
    ws['B' + str(row_counter)] = country_dict['name']
    ws['C' + str(row_counter)] = country_dict['i_1']
    ws['D' + str(row_counter)] = country_dict['i_2']
    ws['E' + str(row_counter)] = country_dict['i_3']
    ws['F' + str(row_counter)] = country_dict['i_4']
    
wb.save(filename ='output.xlsx')